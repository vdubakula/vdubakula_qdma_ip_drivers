import xml.etree.ElementTree as ET
from argparse import ArgumentParser
import openpyxl
import shutil
import re
import sys, os
from array import array
import platform
import traceback

generate_perf_result_desc="Generated qdma gtest results by parsing the output xml files"
usage = "qdma_gtest_result_parser.py -f <xml_file>"
parser = ArgumentParser(prog='qdma_gtest_result_parser.py', usage=usage, description=generate_perf_result_desc, add_help=True)
parser.add_argument("-f", "--in_xml_file",
                    action="store", type=str, dest="xml_file", default = None,
                    help = "xml file of gtest output result")
parser.add_argument("-t", "--tmpl_file",
                    action="store", type=str, dest="tmpl_file",
                    help = "template file for output result")
parser.add_argument("-c", "--con_xml_file",
                    action="store", type=str, dest="consol_file", default=None,
                    help = "Consolidated XML file")

args = parser.parse_args()
xml_file = args.xml_file
tmpl_file = args.tmpl_file
consol_file = args.consol_file

if (xml_file == None):
    print("-f option mandatory")
    sys.exit()
consol_root = None
iteration = None
if not (consol_file == None):
    if not os.path.exists(consol_file):
        consol_root = ET.Element("testsuites")
        consol_tree = ET.ElementTree(consol_root)
        consol_tree.write(consol_file)
    consolidated_tree = ET.parse(consol_file)
    if consolidated_tree == None:
        print("%s: Not an xml file" %consol_file)
        sys.exit()
    consol_root = consolidated_tree.getroot()
    itr_num = len(consol_root.getchildren())
    itr_attrib = {}
    itr_attrib['number'] = str(itr_num)
    iteration = ET.SubElement(consol_root, 'Iteration', itr_attrib)

element_tree = ET.parse(xml_file)
if element_tree == None:
    print("%s: Not an xml file" %xml_file)
    sys.exit()
elem_root = element_tree.getroot()


def strip_word(name, word="DISABLED"):
    namearr = name.split('_')
    if word in namearr:
        namearr.remove(word)
    nname = "_".join(namearr)

    return nname

def get_tg_cell(ws, _tg_name):
    tg_name = strip_word(_tg_name, "VF")
    for row in range(ws.max_row + 1):
        if (ws._get_cell(row + 1, 1).value == tg_name):
                return row + 1
    return -1

def get_tc_cell(ws, lrow, tc_name):
    for row in range(ws.max_row + 1 - lrow):
        tc_row = lrow + row
        if (ws._get_cell(tc_row, 2).value == tc_name):
                return tc_row
    return -1

'''
reorder testcases class-wise
'''
if not (consol_file == None):
    testgroup_dict = {}
    for testsuite in elem_root.findall('testsuite'):
        testsuite_dict = {}
        for testcase in testsuite.findall('testcase'):
            if 'classname' in testcase.attrib.keys():
                classname = testcase.attrib['classname']
                testgroup = 'default'
                tname = ''
                if '/' in classname:
                    arr = classname.split('/')
                    tcstr = strip_word(arr[0])
                    tcarr = tcstr.split('_')
                    testgroup_arr = tcarr
                    testgroup = testgroup_arr[0]
                    if (testgroup == "VF"):
                        testgroup += "_" + testgroup_arr[1]
                    tname = strip_word(tcstr, testgroup)
                    classname = arr[1]
                else:
                    testcasename = testcase.attrib['name']
                    tcarr = (strip_word(testcasename)).split("_")
                    testgroup_arr = tcarr
                    testgroup = testgroup_arr[0]
                    if (testgroup == "VF"):
                        testgroup += "_" + testgroup_arr[1]
                    tname = strip_word(testcasename, testgroup)

                if not (classname in testsuite_dict.keys()):
                    ts_attrib = testsuite.attrib
                    ts_attrib['classname'] = classname
                    ts = ET.SubElement(iteration, 'testsuite', testsuite.attrib)
                    testsuite_dict[classname] = ts
                tgkey = classname + "/" + testgroup
                if not (tgkey in testgroup_dict.keys()):
                    ts = testsuite_dict[classname]
                    tg_attrib = {}
                    tg_attrib['name'] = testgroup
                    tg_attrib['sname'] = tname
                    tg = ET.SubElement(ts, 'testgroup', tg_attrib)
                    testgroup_dict[classname + "/" + testgroup] = tg
                tg = testgroup_dict[classname + "/" + testgroup]
                tc_attrib = testcase.attrib
                tc_attrib['disabled'] = str(tc_attrib['name'].startswith("DISABLED"))
                for failnode in testcase.findall('failure'):
                    tc_attrib['failed'] = "yes"
                tc = ET.SubElement(tg, 'testcase', tc_attrib)

    consolidated_tree.write(consol_file, encoding='utf8', xml_declaration=True, method='xml')
else:
    targetdir = os.path.dirname(xml_file)
    arr = os.path.basename(xml_file).split('.')
    xml_base_filename = arr[0] + ".xlsx"
    if (platform.system() == "Windows"):
        target_file = os.path.join(targetdir, xml_base_filename)
    else :
        target_file = os.path.join(os.sep, targetdir, xml_base_filename)
    base_tmpl_file = os.path.basename(tmpl_file)
    shutil.copy(tmpl_file, target_file)
    if (platform.system() == "Windows"):
        tmpl_file = os.path.join(targetdir, base_tmpl_file)
    else:
        tmpl_file = os.path.join(os.sep, targetdir, base_tmpl_file)
    wb = openpyxl.load_workbook(target_file)
    if wb == None:
        print("%s: No Such template file")
        sys.exit()
    for itr in element_tree.findall('Iteration'):
        itr_num = int((itr.attrib)['number'])
        for ts in itr.findall('testsuite'):
            for tg in ts.findall('testgroup'):
                result = ""
                testcases = tg.findall('testcase')
                num_tc = len(testcases)
                for tc in testcases:
                    tc_status = (tc.attrib)['status']
                    if tc_status == "run":
                        if "failed" in (tc.attrib).keys():
                            result = "FAIL"
                        else:
                            result = "PASS"
                    elif tc_status == "notrun":
                        result = "N/A"
                    else:
                        result = "FAIL"
                    classname= (ts.attrib)['classname']
                    test_cname = strip_word(classname, "qdma")
                    test_cname = strip_word(test_cname, "test")
                    try:
                        if ("VF_" in tc.attrib["classname"]):
                            test_cname += "_vf"
                        elif ("_vf" not in test_cname.lower()):
                            if not ("vf" == test_cname.lower()):
                                test_cname += "_pf"
                        ws = wb[test_cname]
                    except KeyError as e:
                        print("No sheet with name %s" %test_cname)
                        continue
                    col = 1
                    row = get_tg_cell(ws, (tg.attrib)['name'])
                    if (row == -1):
                        print("missing %s_%s test case in template" %(tg.attrib['name'], tg.attrib['sname']))
                        continue
                    col = col + 6
                    if not (num_tc == 1):
                        tgname = strip_word(tc.attrib['classname'])
                        tgname_arr = tgname.split('/')
                        tgclass = tgname_arr[0]
                        tcname = strip_word(strip_word(tgclass, "VF"), strip_word(tg.attrib['name'], "VF")) + "/" + strip_word((tc.attrib)['name'])
                        row = get_tc_cell(ws, row, tcname)
                    if (row == -1):
                        print("missing %s_%s/%s test case in template" %(tg.attrib['name'], tg.attrib['sname'], tc.attrib['name']))
                        continue
                    else:
                        try:
                            ws.cell(row=row, column=(col + itr_num), value=str(result))
                            if (itr_num == 0):
                                ws.cell(row=row, column=(col - 1), value=str(result))
                            elif (result == "FAIL"):
                                ws.cell(row=row, column=(col - 1), value=str(result))
                        except AttributeError as e:
                            print("exception at %u %u in %s %s" %(row, col, test_cname, str(result)))
                            traceback.print_exc()
                            break

    wb.save(target_file)
