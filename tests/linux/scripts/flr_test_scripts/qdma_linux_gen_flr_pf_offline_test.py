import os
import shutil
import glob
import csv
from argparse import ArgumentParser
import re
import openpyxl
from xlrd import open_workbook
from xlrd import XLRDError

wb = None

# This python script generates test report for QDMA Linux Driver FLR+PF Offline Test.
# The script takes in the output file name as input and populates the different fields in the Output Test Report XLS file.
# The Test Report has an Overview sheet and Test Results sheet for each mode.
# The overview sheet has the top level summary of the entire test which is derived from the results
# present in Test Results sheet for each mode.
# For better understanding on how to use the script use "python qdma_linux_gen_flr_test.py -h"

generate_flr_result_desc="Generate Stress Test results by populating entires in the Test Report"
usage = """qdma_linux_gen_flr_test.py -o <output_file> -u <update_type> -q <num_queues> -m <driver mode>
			-n <num_vfs> -v <driver_version> -i <test_case_id> -r <test_result>"""
parser = ArgumentParser(prog='qdma_linux_gen_flr_test.py', usage=usage, description=generate_flr_result_desc, add_help=True)

parser.add_argument("-o", "--ofile",
                    action="store", type=str, dest="output_file",
                    help = "The output report file")

parser.add_argument("-u", "--update_type",
                    action="store", type=str, dest="update_type",
					choices=["Overview","Result"],
                    help = "Overview to update the overview sheet. Result to update the test case execution result for the test case")


#Arguments for populating the Overview Sheet

parser.add_argument("-q", "--queues",
                    action="store", type=str, dest="queues", default="unknown",
                    help = "Number of queues used during the stress test")

parser.add_argument("-n", "--num_vfs",
                    action="store", type=str, dest="num_vfs", default="unknown",
                    help = "Number of VFs")
					
parser.add_argument("-v", "--driver_version",
                    action="store", type=str, dest="driver_version", default="unknown",
                    help = "Driver Version")

parser.add_argument("-b", "--bitstream",
                    action="store", type=str, dest="bitstream", default="unknown",
                    help = "bitstream")
					
parser.add_argument("-d", "--date_of_test",
                    action="store", type=str, dest="date_of_test", default="unknown",
                    help = "date field")
					
parser.add_argument("-t", "--device_type",
                    action="store", type=str, dest="device_type", default="unknown",
                    help = "device_type")
					
#Arguments for populating the Test Results
parser.add_argument("-m", "--mode",
                    action="store", type=str, dest="mode", default="unknown",
                    help = "Mode in which the driver was loaded")
					
parser.add_argument("-i", "--test_case_id",
                    action="store", type=str, dest="test_case_id", default="unknown",
                    help = "Test Case to be updated")

parser.add_argument("-r", "--result",
                    action="store", type=str, dest="test_result", default="unknown",
                    help = "Result of the test case - PASS/FAIL")


					

args = parser.parse_args()
update_type=args.update_type
driver_version=args.driver_version
mode = args.mode
queues = args.queues
test_result = args.test_result
num_vfs = args.num_vfs
outputfile = args.output_file
test_case_id = args.test_case_id
bitstream=args.bitstream
date_of_test=args.date_of_test
device_type=args.device_type
result_sheet_row_base = 2
result_sheet_col_base = 3
overview_poll_mode_row_base = 11
overview_auto_mode_row_base = 12
overview_intr_mode_row_base = 13
overview_aggr_mode_row_base = 14


if (outputfile != None):
	if (update_type == "Overview"):
		wb = openpyxl.load_workbook(outputfile)
		sheet = wb['Overview']
		
		#updating the driver version
		sheet.cell(1,2).value = driver_version
		#updating the driver version
		sheet.cell(2,2).value = date_of_test
		#updating the device type
		sheet.cell(3,2).value = device_type
		#updating the driver version
		sheet.cell(4,2).value = bitstream
		#updating the Number of VFs
		sheet.cell(5,4).value = int(num_vfs)
		#updating the total number of queues
		sheet.cell(6,4).value = int(queues)
	
		wb.save(outputfile)

	elif (update_type == "Result"):
		wb = openpyxl.load_workbook(outputfile)
		if (mode == "Direct_Interrupt"):
			sheet = wb['Direct_Interrupt']
		elif (mode == "Poll"):
			sheet = wb['Poll']
		elif (mode == "Auto"):
			sheet = wb['Auto']
		elif (mode == "Interrupt_Aggr"):
			sheet = wb['Interrupt_Aggr']

		sheet.cell(int(result_sheet_row_base) + int(test_case_id) - 1 ,result_sheet_col_base ,test_result)

		wb.save(outputfile)
	else:
		print(usage)
else:
	print(usage)


