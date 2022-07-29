import os
import shutil
import glob
import csv
from argparse import ArgumentParser
import re
import openpyxl
from xlrd import open_workbook
from xlrd import XLRDError

wb = None

# This python script generates test report for QDMA Linux Driver Stress Test.
# The script takes in the output file name as input and populates the different fields in the Output Test Report XLS file.
# The Test Report has an Overview sheet and Statistics sheet for each mode.
# The overview sheet has the top level summary of the entire test which is derived from the data
# present in Statistics sheet for each mode.
# For better understanding on how to use the script use "python qdma_linux_gen_stress_test.py -h"

generate_perf_result_desc="Generate Stress Test results by populating entires in the Test Report"
usage = """qdma_linux_gen_stress_test.py -o <output_file> -u <update_type> -d <duration in secs> -q <num_queues> -m <driver mode> -p <num_pfs>
			-v <num_vfs> -i <row_to_be_updated> -f <function_id> --st_h2c_num_pkt <ST H2C num packets>
			--mm_h2c_num_pkt <MM H2C num packets > --st_c2h_num_pkt <ST C2H num packets> --mm_c2h_num_pkt <ST C2H num packets>
			--mm_errs <Errors in MM mode> --st_c2h_errs <Errors in ST C2H> --st_h2c_errs <Errors in ST H2C>"""
parser = ArgumentParser(prog='qdma_linux_gen_stress_test.py', usage=usage, description=generate_perf_result_desc, add_help=True)

parser.add_argument("-o", "--ofile",
                    action="store", type=str, dest="output_file",
                    help = "The output report file")

parser.add_argument("-u", "--update_type",
                    action="store", type=str, dest="update_type",
					choices=["Overview","Stat"],
                    help = "Overview to update the overview sheet. Stat to fill data into the statistics sheet for the individual modes")


#Arguments for populating the Overview Sheet
parser.add_argument("-d", "--duration",
                    action="store", type=str, dest="duration", default="unkown",
                    help = "Duration of the stress test in seconds")

parser.add_argument("-q", "--queues",
                    action="store", type=str, dest="queues", default="unkown",
                    help = "Number of queues used during the stress test")

parser.add_argument("-m", "--mode",
                    action="store", type=str, dest="mode", default="unkown",
                    help = "Mode in which the driver was loaded")

parser.add_argument("-p", "--num_pfs",
                    action="store", type=str, dest="num_pfs", default="unkown",
                    help = "Number of PFs")

parser.add_argument("-v", "--num_vfs",
                    action="store", type=str, dest="num_vfs", default="unkown",
                    help = "Number of VFs")


#Arguments for populating the auxillary data sheets
parser.add_argument("-i", "--index",
                    action="store", type=str, dest="index", default="unkown",
                    help = "Row to be updated in the Stat sheet")

parser.add_argument("-f", "--func_id",
                    action="store", type=str, dest="func_id", default="unkown",
                    help = "Function ID")

parser.add_argument("--st_h2c_num_pkt",
                    action="store", type=str, dest="st_h2c_num_pkt", default="unkown",
                    help = "ST H2C Num Pkt")

parser.add_argument("--st_c2h_num_pkt",
                    action="store", type=str, dest="st_c2h_num_pkt", default="unkown",
                    help = "ST C2H Num Pkt")

parser.add_argument("--mm_h2c_num_pkt",
                    action="store", type=str, dest="mm_h2c_num_pkt", default="unkown",
                    help = "MM H2C Num Pkt")

parser.add_argument("--mm_c2h_num_pkt",
                    action="store", type=str, dest="mm_c2h_num_pkt", default="unkown",
                    help = "ST C2H Num Pkt")

parser.add_argument("--mm_errs",
                    action="store", type=str, dest="mm_errs", default="0",
                    help = "mm errors")

parser.add_argument("--st_c2h_errs",
                    action="store", type=str, dest="st_c2h_errs", default="0",
                    help = "ST C2H errs")

parser.add_argument("--st_h2c_errs",
                    action="store", type=str, dest="st_h2c_errs", default="0",
                    help = "ST H2C errs")

args = parser.parse_args()
update_type=args.update_type
duration = args.duration
mode = args.mode
queues = args.queues
num_pfs = args.num_pfs
num_vfs = args.num_vfs
outputfile = args.output_file
index = args.index
st_h2c_num_pkt = args.st_h2c_num_pkt
st_c2h_num_pkt = args.st_c2h_num_pkt
mm_h2c_num_pkt = args.mm_h2c_num_pkt
mm_c2h_num_pkt = args.mm_c2h_num_pkt
mm_errs=args.mm_errs
st_c2h_errs=args.st_c2h_errs
st_h2c_errs=args.st_h2c_errs
func_id = args.func_id
stat_sheet_row_base = 2
stat_sheet_col_base = 2
overview_poll_mode_row_base = 13
overview_auto_mode_row_base = 14
overview_intr_mode_row_base = 15
overview_aggr_mode_row_base = 16


if (outputfile != None):
	if (update_type == "Overview"):
		wb = openpyxl.load_workbook(outputfile)
		sheet = wb['Overview']
		#updating the duration
		sheet.cell(7,4).value = int(duration)
		#updating the Number of PFs
		sheet.cell(8,4).value = int(num_pfs)
		#updating the Number of VFs
		sheet.cell(9,4).value = int(num_vfs)
		#updating the total number of queues
		sheet.cell(10,4).value = int(queues)

		if (mode == "poll"):
			sheet.cell(overview_poll_mode_row_base,8).value = int(mm_errs)
			sheet.cell(overview_poll_mode_row_base,9).value = int(st_h2c_errs)
			sheet.cell(overview_poll_mode_row_base,10).value = int(st_h2c_errs)
		elif (mode == "auto"):
			sheet.cell(overview_auto_mode_row_base,8).value = int(mm_errs)
			sheet.cell(overview_auto_mode_row_base,9).value = int(st_h2c_errs)
			sheet.cell(overview_auto_mode_row_base,10).value = int(st_c2h_errs)
		elif (mode == "intr"):
			sheet.cell(overview_intr_mode_row_base,8).value = int(mm_errs)
			sheet.cell(overview_intr_mode_row_base,9).value = int(st_h2c_errs)
			sheet.cell(overview_intr_mode_row_base,10).value = int(st_c2h_errs)
		elif (mode == "aggr"):
			sheet.cell(overview_aggr_mode_row_base,8).value = int(mm_errs)
			sheet.cell(overview_aggr_mode_row_base,9).value = int(st_h2c_errs)
			sheet.cell(overview_aggr_mode_row_base,10).value = int(st_c2h_errs)

		wb.save(outputfile)

	elif (update_type == "Stat"):
		wb = openpyxl.load_workbook(outputfile)
		if (mode == "intr"):
			sheet = wb['Direct_Intr_Mode_Stats']
		elif (mode == "poll"):
			sheet = wb['Poll_Mode_Stats']
		elif (mode == "auto"):
			sheet = wb['Auto_Mode_Stats']
		elif (mode == "aggr"):
			sheet = wb['Indirect_Intr_Mode_Stats']

		sheet.cell(int(stat_sheet_row_base) + int(index) ,stat_sheet_col_base ,func_id)
		sheet.cell(int(stat_sheet_row_base) + int(index) ,stat_sheet_col_base + 1).value = int(mm_h2c_num_pkt)
		sheet.cell(int(stat_sheet_row_base) + int(index) ,stat_sheet_col_base + 2).value = int(mm_c2h_num_pkt)
		sheet.cell(int(stat_sheet_row_base) + int(index) ,stat_sheet_col_base + 3).value = int(st_h2c_num_pkt)
		sheet.cell(int(stat_sheet_row_base) + int(index) ,stat_sheet_col_base + 4).value = int(st_c2h_num_pkt)

		wb.save(outputfile)
	else:
		print(usage)
else:
	print(usage)


