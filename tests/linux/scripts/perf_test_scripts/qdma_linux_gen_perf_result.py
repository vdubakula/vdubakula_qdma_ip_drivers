import os
import shutil
import glob
import csv
from argparse import ArgumentParser
import re
import openpyxl
from xlrd import open_workbook
from xlrd import XLRDError

generate_perf_result_desc="Generated performance results by parsing the output files"
usage = "qdma_linux_gen_perf_result.py -d input_directory -t template_dir"
parser = ArgumentParser(prog='qdma_linux_gen_perf_result.py', usage=usage, description=generate_perf_result_desc, add_help=True)
parser.add_argument("-d", "--dir",
                    action="store", type=str, dest="rootdir",
                    help = "Directory containing the output log files from dmaperf")
parser.add_argument("-t", "--tdir",
                    action="store", type=str, dest="templdir",
                    help = "Directory containing the templates of output result")
parser.add_argument("-o", "--out_file_suffix",
                    action="store", type=str, dest="out_suffix", default="unkown",
                    help = "Suffix that needs to be added to the out file")
args = parser.parse_args()
rootdir = args.rootdir
template_dir=args.templdir
target = ".." + str(os.sep)
resultdir = str(os.path.join(os.sep, rootdir, target))
out_file_suffix=args.out_suffix

row_init = 4
col_init = 0
rd_col_offset = 3
pps_col_offset = 1
bw_col_offset = 0
cpu_col_offset = 2
parser_version = 2
rd_regx = r"[A-Z:\s\w\d=./]*READ:\s*total\s*pps\s*=\s*(?P<rd_pps>\d+)\s*BW\s*=\s*(?P<rd_bw>[\d.\s\w/]+)$"
wr_regx = r"[\s\w\d=./]*WRITE:\s*total\s*pps\s*=\s*(?P<wr_pps>\d+)\s*BW\s*=\s*(?P<wr_bw>[\d.\s\w/]+)$"
sar_regx = r"[\d:\w\s]+all\s+(?P<usr_sar>[\d.]+)\s+[\d.]+\s+(?P<sys_sar>[\d.]+)[\d\s.]+$"
perf_rd_regx = re.compile(rd_regx, re.MULTILINE)
perf_wr_regx = re.compile(wr_regx, re.MULTILINE)
perf_sar_regx = re.compile(sar_regx, re.MULTILINE)

'''
get template file
'''
def get_template_file(mode):
    basedir = os.path.basename(rootdir)
    tmpl_file_name = "perf_result_template_" + mode + "_" + basedir + ".xlsx"
    return tmpl_file_name
                
def get_file_dir(filename):
    arr = str(filename).split('.') 
    basefilename = arr[0]
    arr = str(basefilename).split('_')
    if (parser_version == 1):
        dir = arr[2]
        mode = "st"
    else:
        dir = arr[3]
        
    return dir

'''
get row index of corresponding io size
'''
def get_io_cell(file):
    filename = os.path.basename(file)
    dirname = os.path.basename(os.path.dirname(file))
    if (not (str(filename).startswith("log_") or str(filename).startswith("sar_"))):
        return None, 0, 0
    if (filename == "sar_1536_c2h.txt"):
        filename = "sar_1536_c2h.txt"
    arr = str(filename).split('.') 
    basefilename = arr[0]
    arr = str(basefilename).split('_')
    io_sz = arr[1]
    mode = arr[2]
    dir = get_file_dir(filename)
    row = row_init
    col = col_init
    if (dir == "c2h"):
        col = 4
    else:
        col = 1
    tmpl_file_name = get_template_file(mode)
    tmpl_file = os.path.join(os.sep, template_dir, tmpl_file_name)
    wb = open_workbook(tmpl_file)
    scan_io_sz = False
    sheetdir = "unidir"
    if (dir == "bi"):
        sheetdir = "bidir"
    sheetname = "perf_result_" + sheetdir
    if (mode == "st"):
        arr = str(dirname).split('_')
        cfg_name = arr[3]
        try:
            sheet = wb.sheet_by_name(cfg_name)
            sheetname = cfg_name
        except XLRDError:
            print("could not find sheet %s\n" %cfg_name)
            return None, -1, -1
        if (cfg_name == str(sheet.name)): #search the sheet
            for lrow in range(sheet.nrows):
                if (str(sheet.cell(lrow, 0).value) == str(dirname)): #search cfg_name
                    row = lrow + 2
                    scan_io_sz = True
                if (scan_io_sz == True):
                    cellvalstr = str(sheet.cell(row, 0).value)
                    cellval = ""
                    if not (cellvalstr.strip() == ""):
                        try:
                            cellval = str(int(float(cellvalstr)))
                        except ValueError as err:
                            print(str(err))
                    if (cellval == io_sz): #search the io_sz
                        break
                    row = row + 1
    else:
        try:
            sheet = wb.sheet_by_name(sheetname)
        except XLRDError:
            print("could not find sheet %s\n" %sheetname)
            return None, -1, -1
        for lrow in range(sheet.nrows):
            if (str(sheet.cell(lrow, 0).value) == str(dirname)): #search cfg_name
                row = lrow + 2
                scan_io_sz = True
            if (scan_io_sz == True):
                cellvalstr = str(sheet.cell(row, 0).value)
                cellval = ""
                if not (cellvalstr.strip() == ""):
                    cellval = str(int(float(cellvalstr)))
                if (cellval == io_sz): #search the io_sz
                    break
                row = row + 1
    
    if (scan_io_sz == False):
        row = -1
        sheetname = None
        
    return sheetname, row, col


def toMb(value, unit):
    val = float(value)
    unit = unit.strip()
    bw = 0
    if (unit == "KB/sec"):
        bw = ((val*8)/1000)
    elif (unit == "MB/sec"):
        bw = (val*8)
    elif (unit == "GB/sec"):
        bw = (val*1000*8)
    else:
       bw = ((val*8)/1000000)
    
    return bw


'''
parse log files
'''
def parse_log_file(filedata, param):
    pps = 0
    bw = 0
    bw_str = ""
    if (param == "rd"):
        perf_rd = perf_rd_regx.match(filedata)
        if (perf_rd == None):
            return 0, 0
        pps = int(perf_rd.group('rd_pps'))
        bw_str = perf_rd.group('rd_bw')
    else:
        perf_wr = perf_wr_regx.match(filedata)
        if (perf_wr == None):
            return 0, 0
        pps = int(perf_wr.group('wr_pps'))
        bw_str = perf_wr.group('wr_bw')
        
    arr = bw_str.split(' ')
    value = arr[0]
    units = arr[1]
    bw = toMb(value, units)
    
    return pps, bw


'''
parse sar files
'''
def parse_sar_file(filedata):
    cpu = 0
    
    sar = perf_sar_regx.match(filedata)
    if (sar == None):
            return 0
    sys_cpu = float(sar.group('sys_sar'))
    usr_cpu = float(sar.group('usr_sar'))
    cpu = float(sys_cpu + usr_cpu)
    
    return cpu


'''
process output logs and write to xlsx file
'''
excel_dict = {}
wb = None
out_path = ""
for root, subdirs, files in os.walk(rootdir):
    for folder in subdirs:
        print("processing %s" %folder)
        arr = str(folder).split('_')
        cfg_mode = arr[0]
        tmpl_file_name = get_template_file(cfg_mode)
        if not (tmpl_file_name in excel_dict.keys()):
            excel_dict[tmpl_file_name] = 1
            tmpl_file = os.path.join(os.sep, template_dir, tmpl_file_name)
            basedir = os.path.basename(rootdir)
            out_path = os.path.join(os.sep, resultdir, "qdma_perf_result_" + out_file_suffix + "_" + basedir + ".xlsx")
            shutil.copy(tmpl_file, out_path)
            wb = openpyxl.load_workbook(tmpl_file)
        folder_path = os.path.join(os.sep, rootdir, folder)
        for direc, subdir, files in os.walk(folder_path):
            for file in files:
                file_path = os.path.join(os.sep, rootdir, folder, file)
                sheetname, row, col = get_io_cell(file_path)
                '''
                add 1 to row and col for openpyxl
                '''
                row = row + 1
                col = col + 1
                if (sheetname == None):
                    continue
                sheet = wb[sheetname]
                filename = file
                arr = str(filename).split('_')
                dir = get_file_dir(filename)
                inputdata = ""
                try:
                    f = open(file_path,"r")
                    input =f.readlines()
                    f.close()
                    in_file_lines = len(input)
                    for i in range(in_file_lines):
                        if (filename.startswith("log_")):
                            if ((input[i].startswith("READ")) or (input[i].startswith("WRITE"))):
                                inputdata+= input[i]
                        else:
                            if ("all" in input[i].split(' ')):
                                inputdata+= input[i]
                                break
                except IOError as err:
                    print("something went wrong trying to read input file %s: %s\n" % (filename,err))
                if (str(file).startswith("log")):
                    if (dir == "bi"):
                        pps, bw = parse_log_file(inputdata, "wr")
                        sheet.cell(row, col + bw_col_offset, bw)
                        sheet.cell(row, col + pps_col_offset, pps)
                        col = col + rd_col_offset
                        pps, bw = parse_log_file(inputdata, "rd")
                        sheet.cell(row, col + bw_col_offset, bw)
                        sheet.cell(row, col + pps_col_offset, pps)
                    elif (dir == "h2c"):
                        pps, bw = parse_log_file(inputdata, "wr")
                        sheet.cell(row, col + bw_col_offset, bw)
                        sheet.cell(row, col + pps_col_offset, pps)
                    else:
                        pps, bw = parse_log_file(inputdata, "rd")
                        sheet.cell(row, col + bw_col_offset, bw)
                        sheet.cell(row, col + pps_col_offset, pps)
                elif (str(file).startswith("sar")):
                    cpu = parse_sar_file(inputdata)
                    sheet.cell(row, col + cpu_col_offset, cpu)
                else:
                    print("Unknown file")
                    continue
        wb.save(out_path)
