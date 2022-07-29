import os
import shutil
import glob
import csv
from argparse import ArgumentParser
import re
import openpyxl
from xlrd import open_workbook
from xlrd import XLRDError

generate_latency_result_desc="Generated ping pong latency results by parsing the output files"
usage = "qdma_linux_gen_latency_result.py -d input_directory -t template_dir"
parser = ArgumentParser(prog='qdma_linux_gen_latency_result.py', usage=usage, description=generate_latency_result_desc, add_help=True)
parser.add_argument("-d", "--dir",
                    action="store", type=str, dest="rootdir",
                    help = "Directory containing the output log files from dmalat")
parser.add_argument("-t", "--tdir",
                    action="store", type=str, dest="templdir",
                    help = "Directory containing the templates of output result")
parser.add_argument("-o", "--out_file_suffix",
                    action="store", type=str, dest="out_suffix", default="unkown",
                    help = "Suffix that needs to be added to the out file")
					
parser.add_argument("-s", "--cpu_speed",
                    action="store", type=str, dest="CPU_speed", default="unkown",
                    help = "CPU Speed in MHz")
args = parser.parse_args()
rootdir = args.rootdir
template_dir=args.templdir
cpu_speed=args.CPU_speed
target = ".." + str(os.sep)
resultdir = str(os.path.join(os.sep, rootdir, target))
out_file_suffix=args.out_suffix

row_init = 4
col_init = 0
rd_col_offset = 3
pps_col_offset = 1
bw_col_offset = 0
cpu_col_offset = 2
parser_version = 2

min_regx = r"[A-Z:\s\w\d=./]*Min\s*Ping\s*Pong\s*Latency\s*=\s*(?P<min_latency>\d+)$"
max_regx = r"[A-Z:\s\w\d=./]*Max\s*Ping\s*Pong\s*Latency\s*=\s*(?P<max_latency>\d+)$"
avg_regx = r"[A-Z:\s\w\d=./]*Avg\s*Ping\s*Pong\s*Latency\s*=\s*(?P<avg_latency>\d+)$"

min_lat_regx = re.compile(min_regx, re.MULTILINE)
max_lat_regx = re.compile(max_regx, re.MULTILINE)
avg_lat_regx = re.compile(avg_regx, re.MULTILINE)

'''
get template file
'''
def get_template_file():
    basedir = os.path.basename(rootdir)
    tmpl_file_name = "ping_pong_latency_result_template.xlsx"
    return tmpl_file_name
                
def get_file_dir(filename):
    arr = str(filename).split('.') 
    basefilename = arr[0]
    arr = str(basefilename).split('_')
    if (parser_version == 1):
        dir = arr[2]
        mode = "st"
    else:
        dir = arr[3]
        
    return dir

'''
get row index of corresponding io size
'''
def get_io_cell(file):
	filename = os.path.basename(file)
	dirname = os.path.basename(os.path.dirname(file))
	if (not (str(filename).startswith("log_") or str(filename).startswith("sar_"))):
		return None, 0, 0

	arr = str(filename).split('.') 
	basefilename = arr[0]
	arr = str(basefilename).split('_')
	io_sz = arr[1]
	mode = arr[2]
	dir = get_file_dir(filename)
	row = row_init
	col = 1
	tmpl_file_name = get_template_file()
	tmpl_file = os.path.join(os.sep, template_dir, tmpl_file_name)
	wb = open_workbook(tmpl_file)
	scan_io_sz = False
	sheetdir = "bidir"
	sheetname = "Ping_Pong_Latency_Results"
    
	arr = str(dirname).split('_')
	cfg_name = arr[3]
	try:
		sheet = wb.sheet_by_name(sheetname)
	except XLRDError:
		print("could not find sheet %s\n" %sheetname)
		return None, -1, -1
		
	
	for lrow in range(sheet.nrows):
		if (str(sheet.cell(lrow, 0).value) == str(dirname)): #search cfg_name
			row = lrow + 2
			scan_io_sz = True
		if (scan_io_sz == True):
			cellvalstr = str(sheet.cell(row, 0).value)
			cellval = ""
			if not (cellvalstr.strip() == ""):
				try:
					cellval = str(int(float(cellvalstr)))
				except ValueError as err:
					print(str(err))
			if (cellval == io_sz): #search the io_sz
				break
			row = row + 1
	
	if (scan_io_sz == False):
		row = -1
		sheetname = None
		
	return sheetname, row, col


'''
parse log files
'''
def parse_log_file(filedata, param):
	minimum_latency = 0
	maximum_latency = 0
	average_latency = 0
	if (param == "min_lat"):
		min_lat = min_lat_regx.match(filedata)
		if (min_lat == None):
			return 0, 0, 0
		minimum_latency = int(min_lat.group('min_latency'))
	elif (param == "max_lat"):
		max_lat = max_lat_regx.match(filedata)
		if (max_lat == None):
			return 0, 0, 0
		maximum_latency = int(max_lat.group('max_latency'))
	elif (param == "avg_lat"):
		avg_lat = avg_lat_regx.match(filedata)
		if (avg_lat == None):
			return 0, 0, 0
		average_latency = int(avg_lat.group('avg_latency'))

	return minimum_latency, maximum_latency, average_latency



'''
process output logs and write to xlsx file
'''
excel_dict = {}
wb = None
out_path = ""
for root, subdirs, files in os.walk(rootdir):
    for folder in subdirs:
        print("processing %s" %folder)
        arr = str(folder).split('_')
        tmpl_file_name = get_template_file()
        if not (tmpl_file_name in excel_dict.keys()):
            excel_dict[tmpl_file_name] = 1
            tmpl_file = os.path.join(os.sep, template_dir, tmpl_file_name)
            basedir = os.path.basename(rootdir)
            out_path = os.path.join(os.sep, resultdir, "qdma_ping_pong_latency_result_" + out_file_suffix + "_" + basedir + ".xlsx")
            shutil.copy(tmpl_file, out_path)
            wb = openpyxl.load_workbook(tmpl_file)
        folder_path = os.path.join(os.sep, rootdir, folder)
        for direc, subdir, files in os.walk(folder_path):
            for file in files:
                file_path = os.path.join(os.sep, rootdir, folder, file)
                sheetname, row, col = get_io_cell(file_path)
                '''
                add 1 to row and col for openpyxl
                '''
                row = row + 1
                col = col + 1
                if (sheetname == None):
                    continue
                sheet = wb[sheetname]
                filename = file
                arr = str(filename).split('_')
                dir = get_file_dir(filename)
                inputdata = ""
                try:
                    f = open(file_path,"r")
                    input =f.readlines()
                    f.close()
                    in_file_lines = len(input)
                    for i in range(in_file_lines):
                        if (filename.startswith("log_")):
                            if ((input[i].startswith("Min Ping Pong")) or (input[i].startswith("Max Ping Pong") or (input[i].startswith("Avg Ping Pong")))):
                                inputdata+= input[i]
                        else:
                            if ("all" in input[i].split(' ')):
                                inputdata+= input[i]
                                break
                except IOError as err:
                    print("something went wrong trying to read input file %s: %s\n" % (filename,err))
                if (str(file).startswith("log")):
                    if (dir == "bi"):
						min_pp_lat, max_pp_lat, avg_pp_lat = parse_log_file(inputdata, "min_lat")
						sheet.cell(row, 2, ((float)(min_pp_lat)/((float)(cpu_speed))))

						min_pp_lat, max_pp_lat, avg_pp_lat = parse_log_file(inputdata, "max_lat")
						sheet.cell(row, 3, ((float)(max_pp_lat)/((float)(cpu_speed))))

						min_pp_lat, max_pp_lat, avg_pp_lat = parse_log_file(inputdata, "avg_lat")
						sheet.cell(row, 4, ((float)(avg_pp_lat)/((float)(cpu_speed))))

                    continue		
        wb.save(out_path)
