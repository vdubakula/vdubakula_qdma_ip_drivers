import sys
import os
import getopt
import zipfile
import openpyxl
import shutil
from shutil import copyfile
from xlrd import open_workbook
from datetime import datetime
import xml.etree.ElementTree as ET


templateFileName = ""
configFileName = ""
perfExe = r""
linkMode = r""

#-------------------------------------------------------------------
# CONSTANTS
#
testOutputDir   = "PerfResults"
blockSizeOffset = 3 # Offset after test case name
H2CBwOffset     = 2
H2CPpsOffset    = 3
H2CCpuOffset    = 4
C2HBwOffset     = 5
C2HPpsOffset    = 6
C2HCpuOffset    = 7
#-------------------------------------------------------------------

def usage():
    print ("usage : python %s [OPTIONS]\n" % sys.argv[0])
    print ("Executes performance run\n\n")
    print ("Mandatory arguments")
    print ("\t -t, --template_file\t\t Template file with full path (default : x64)")
    print ("\t -c, --config_file\t configuration file for the test")
    print ("\t -e, --executable\t\t diskspd exe full path")
    print ("\n")
    sys.stdout.flush()

class testManager:
    def __init__(self):
        self.threads = "-t1"
        self.outstanding = "-o1"
        self.duration = "-d20"
        self.trigMode = "-Qm1"
        self.pfchEnabled = " "
        self.ringIndex = "-Qr0"
        self.counterIndex = "-Qc0"
        self.timerIndex = "-Qt0"
        self.wbAcc = "-Qw1"
        self.queueDirection = "h2c"
        self.iterations = 1
        self.cmptSize = 0

    def removeComment(self, string):
        string = string.split('#')
        return string[0]

    def parseConfigFile(self, fileName):
        f = open(fileName, "r")
        for line in f:
            line = self.removeComment(line)
            l1, l2 = line.split('=')
            l2 = l2.strip()
            if (l1 == "iterations") :
                self.iterations = int(l2)
            elif (l1 == "queue_dir") :
                self.queueDirection = l2
            elif (l1 == "wb_acc") :
                self.wbAcc = " -Qw"+l2
            elif (l1 == "ring_idx") :
                self.ringIndex = " -Qr"+l2
            elif (l1 == "tmr_idx") :
                self.timerIndex = " -Qt"+l2
            elif (l1 == "cntr_idx") :
                self.counterIndex = " -Qc"+l2
            elif (l1 == "pfch_en") :
                if (l2 == '1') :
                    self.pfchEnabled = " -Qp"
            elif (l1 == "trig_mode") :
                self.trigMode = " -Qm"+l2
            elif (l1 == "run_time") :
                self.duration = " -d"+l2
            elif (l1 == "num_threads") :
                self.threads = " -t"+l2
            elif (l1 == "outstanding") :
                self.outstanding = " -o"+l2
            else :
                print("Invalid configuration option")
                exit(-1)

    def parse_log_file(self, fileName, Dir):
        bw = 0.0
        pps = 0.0
        bbw = 0.0
        bpps = 0.0
        tsec = 0.0

        perfList = []

        tree = ET.parse(fileName)
        root = tree.getroot()

        time_span = root.find('TimeSpan')
        tsec = float(time_span.find('TestTimeSeconds').text)

        cpu_util = time_span.find('CpuUtilization')
        avg_cpu = cpu_util.find('Average')
        cpu = float(avg_cpu.find('UsagePercent').text)

        for ts in time_span.findall('Thread') :
            for targets in ts :
                for ta in targets :
                    if (Dir == 'unidir') :
                        if ta.tag == 'BytesCount' :
                            bw += float(ta.text)
                        if ta.tag == 'IOCount' :
                            pps += float(ta.text)
                    else :
                        if ta.tag == 'WriteBytes' :
                            bw += float(ta.text)
                        if ta.tag == 'WriteCount' :
                            pps += float(ta.text)
                        if ta.tag == 'ReadBytes' :
                            bbw += float(ta.text)
                        if ta.tag == 'ReadCount' :
                            bpps += float(ta.text)

        bw = (bw * 8)/tsec/1000000 #Mb/sec
        pps = pps/tsec

        perfList.append((pps, bw, cpu))
        if (Dir == 'bidir'):
            bbw = (bbw * 8)/tsec/1000000 #Mb/sec
            bpps = bpps/tsec
            perfList.append((bpps, bbw, cpu))

        return perfList

    def executeTest(self, blockSize, testDir, dir, qdir, mode, testNumber, queues, cmptSz, pfetch = None):
        logFileName = testDir + "\\"+"log_" + str(int(blockSize)) + "_" + dir + "_" + qdir + "_" + mode + "_" + str(queues) + ".xml"

        writePercent = ""
        if (dir == 'bidir') :
            writePercent = ' -w50'
        elif (dir == 'unidir'):
            if (qdir == "h2c") :
                writePercent = ' -w100'
            elif ( qdir == "c2h"):
                writePercent = ' -w0'
        
        cmd = perfExe + ' -b' + str(int(blockSize))
        cmd += self.duration + self.threads + self.outstanding + writePercent
        cmd += self.ringIndex + self.trigMode

        if (mode == 'st') :
            cmd += self.timerIndex + self.counterIndex + self.wbAcc + ' -Qo1' + ' -Qz' + cmptSz
            if (pfetch) :
                cmd += ' -Qp'
        else :
            cmd += ' -Qo0'

        for q in range(int(queues)) :
            cmd += ' -q' + str(q)

        cmd += ' -Rxml'
        cmd += ' > ' + logFileName
        print (cmd)
        os.system(cmd)
        return self.parse_log_file(logFileName, dir)
        #perfList = [self.parse_log_file(logFileName, qdir)]
        #return perfList
        #os.startfile(cmd)
        #print testDir, dir, mode, testNumber, queues

def getTestDirection(fileName):
    if ((fileName.find('unidir') != -1)) :
        return 'unidir'
    elif ((fileName.find('bidir') != -1)) :
        return 'bidir'
    else :
        return None

def getTestConfig(testName):
    #print testName
    '''
    Test Config Format
    <mm|st>_<#TestNumber>_<#NumberOfQueues>_cmptsz<#Size>_pfetch
    '''
    if ((testName.find('mm') != -1)) :
        testSplit = testName.split('_')
        mode = testSplit[0]
        testNumber = testSplit[1]
        queues = testSplit[2]
        return mode, testNumber, queues, None, None
    if ((testName.find('st') != -1)) :
        testSplit = testName.split('_')
        mode = testSplit[0]
        testNumber = testSplit[1]
        queues = testSplit[2]
        cmptsz = testSplit[3][6]
        pfetch = False
        if (len(testSplit) == 5) :
            pfetch = True
        return mode, testNumber, queues, cmptsz, pfetch
    else :
        return None, None, None, None, None

def zipdir(path, zipName):
    # ziph is zipfile handle
    zhandle = zipfile.ZipFile(zipName, 'w', zipfile.ZIP_DEFLATED)
    for root, dirs, files in os.walk(path):
        for file in files:
            zhandle.write(os.path.join(root, file))
    zhandle.close()

if __name__ == "__main__":
    try :
        opts, args = getopt.getopt(sys.argv[1:], "t:c:e:l:h", ['templateFileName=', 'configFileName=', 'perfExe=', 'link-mode=', 'help'])
    except getopt.GetoptError:
        usage()
        sys.exit(1)

    # Retrieve the CLI arguments
    for opt, arg in opts:
        if opt in ('-h', '--help'):
            usage()
            sys.exit(0)
        elif opt in ('-t', '--template_file'):
            templateFileName=arg
        elif opt in ('-c', '--config_file'):
            configFileName = arg
        elif opt in ('-e', '--executable'):
            perfExe = arg
        elif opt in ('-l', '--link-mode'):
            linkMode = arg
        else:
            usage()
            sys.exit(1)

    print ("--------------------------------------------------------")
    print ("Running performance script: ")
    print ("Template file \t\t: ", templateFileName)
    print ("Config file \t\t: ", configFileName)
    print ("--------------------------------------------------------")

    try :
        dir = getTestDirection(templateFileName)
        if (None == dir):
            exit("Could not get the right direction from template")

        testOutputDir = testOutputDir + "_" + dir + "_" + linkMode
        if not os.path.exists(testOutputDir):
            os.mkdir(testOutputDir)

        wb = open_workbook(templateFileName)
        wb.release_resources()
    except :
        exit("Could not get the right template file")

    testConf = testManager()
    testConf.parseConfigFile(configFileName)
    
    for iter in range(testConf.iterations) :
        testDir = testOutputDir + '\\' + str(iter)
        
        if not os.path.exists(testDir) :
            os.mkdir(testDir)
        
        wbTemplateFile = testDir+'\\'+os.path.basename(templateFileName)
        copyfile(templateFileName, wbTemplateFile)
        rb = open_workbook(wbTemplateFile)  # Read only workbook
        wb = openpyxl.load_workbook(wbTemplateFile) # write only workbook

        sheetNames = rb.sheet_names()
        try : 
            for sheetName in sheetNames:
                print (sheetName)
                sheet = rb.sheet_by_name(sheetName)
                wsheet = wb[sheetName]
    
                for row in range(sheet.nrows):
                    if (sheet.cell(row, 0).ctype == 1): # Unicode string
                        mode, testNumber, queues, cmptSz, pfetch = getTestConfig(sheet.cell(row, 0).value)
                        if (mode):
                            blockRow = row + blockSizeOffset
    
                            while (blockRow < sheet.nrows):
                                if (sheet.cell(blockRow, 0).ctype == 2) : # Float 
                                    blockSize = sheet.cell(blockRow, 0).value
                                    if (dir != 'bidir') :
                                        if (testConf.queueDirection == 'both') :
                                            if (cmptSz == None or cmptSz == '0') :
                                                perfList = testConf.executeTest(blockSize, testDir, dir, 'h2c', mode, testNumber, queues, cmptSz, pfetch)
                                                pps, bw, cpu = perfList[0]
                                                wsheet.cell(blockRow+1, H2CBwOffset, bw)
                                                wsheet.cell(blockRow+1, H2CPpsOffset, pps)
                                                wsheet.cell(blockRow+1, H2CCpuOffset, cpu)
                                            
                                            perfList = testConf.executeTest(blockSize, testDir, dir, 'c2h', mode, testNumber, queues, cmptSz, pfetch)
                                            pps, bw, cpu = perfList[0]
                                            wsheet.cell(blockRow+1, C2HBwOffset,  bw)
                                            wsheet.cell(blockRow+1, C2HPpsOffset, pps)
                                            wsheet.cell(blockRow+1, C2HCpuOffset, cpu)
                                        elif (testConf.queueDirection == 'h2c'):
                                            perfList = testConf.executeTest(blockSize, testDir, dir, 'h2c', mode, testNumber, queues, cmptSz, pfetch)
                                            pps, bw, cpu = perfList[0]
                                            wsheet.cell(blockRow+1, H2CBwOffset,  bw)
                                            wsheet.cell(blockRow+1, H2CPpsOffset, pps)
                                            wsheet.cell(blockRow+1, H2CCpuOffset, cpu)
                                        elif (testConf.queueDirection == 'c2h'):
                                            perfList = testConf.executeTest(blockSize, testDir, dir, 'c2h', mode, testNumber, queues, cmptSz, pfetch)
                                            pps, bw, cpu = perfList[0]
                                            wsheet.cell(blockRow+1, C2HBwOffset,  bw)
                                            wsheet.cell(blockRow+1, C2HPpsOffset, pps)
                                            wsheet.cell(blockRow+1, C2HCpuOffset, cpu)
                                    else :
                                        perfList = testConf.executeTest(blockSize, testDir, dir, 'bi', mode, testNumber, queues, cmptSz, pfetch)
                                        pps, bw, cpu = perfList[0]
                                        wsheet.cell(blockRow+1, H2CBwOffset,  bw)
                                        wsheet.cell(blockRow+1, H2CPpsOffset, pps)
                                        wsheet.cell(blockRow+1, H2CCpuOffset, cpu)
                                        pps, bw, cpu = perfList[1]
                                        wsheet.cell(blockRow+1, C2HBwOffset,  bw)
                                        wsheet.cell(blockRow+1, C2HPpsOffset, pps)
                                        wsheet.cell(blockRow+1, C2HCpuOffset, cpu)
                                else:
                                    row += blockRow
                                    break
                                blockRow += 1
            wb.save(wbTemplateFile)
        except :
            wb.save(wbTemplateFile)
        #zipName = os.path.basename(templateFileName) + ".zip"
        now = datetime.now()
        zipName = testOutputDir + now.strftime("%H%M%S") + ".zip"
        print ("Zipping directory %s to %s" % (testOutputDir, zipName))
        zipdir(testOutputDir, zipName)
        #os.rename(testOutputDir +  "_" + linkMode, testOutputDir +  "_" + linkMode + now.strftime("%H%M%S"))
        shutil.rmtree(testOutputDir)