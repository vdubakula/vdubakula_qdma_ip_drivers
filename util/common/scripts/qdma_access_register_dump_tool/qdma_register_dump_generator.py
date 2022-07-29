import os
import shutil
import glob
import csv
from argparse import ArgumentParser
import re
import openpyxl
from xlrd import open_workbook
from xlrd import XLRDError
from difflib import SequenceMatcher
import sys

wb = None
REGISTER_NAME_ROW_BASE = 28
REGISTER_NAME_COL = 1
REGISTER_ADDRESS_COL = 2
BITFIELD_NAME_COL = 3
BITFIELD_DEFINITION_COL = 6
DEBUG_MODE_COL=8
FEATURE_COL=29
PF_VF_COL=30

template_copyright = """/*
 * Copyright(c) 2019-2020 Xilinx, Inc. All rights reserved.
 */
"""


template_base_defines = """
#ifdef __cplusplus
extern "C" {
#endif

#include "qdma_platform.h"

#ifdef CHAR_BIT
#undef CHAR_BIT
#endif
#define CHAR_BIT 8

#ifdef BIT
#undef BIT
#endif
#define BIT(n)                  (1u << (n))

#ifdef BITS_PER_BYTE
#undef BITS_PER_BYTE
#endif
#define BITS_PER_BYTE           CHAR_BIT

#ifdef BITS_PER_LONG
#undef BITS_PER_LONG
#endif
#define BITS_PER_LONG           (sizeof(uint32_t) * BITS_PER_BYTE)

#ifdef BITS_PER_LONG_LONG
#undef BITS_PER_LONG_LONG
#endif
#define BITS_PER_LONG_LONG      (sizeof(uint64_t) * BITS_PER_BYTE)

#ifdef GENMASK
#undef GENMASK
#endif
#define GENMASK(h, l) \\
	((0xFFFFFFFF << (l)) & (0xFFFFFFFF >> (BITS_PER_LONG - 1 - (h))))

#ifdef GENMASK_ULL
#undef GENMASK_ULL
#endif
#define GENMASK_ULL(h, l) \\
	((0xFFFFFFFFFFFFFFFF << (l)) & \\
			(0xFFFFFFFFFFFFFFFF >> (BITS_PER_LONG_LONG - 1 - (h))))

#define DEBGFS_LINE_SZ			(81)

#ifdef ARRAY_SIZE
#undef ARRAY_SIZE
#endif
#define ARRAY_SIZE(arr) (sizeof(arr) / \\
							sizeof(arr[0]))


"""
template_footer = """
#ifdef __cplusplus
}
#endif

#endif

"""

rv_check_str= """
	if (rv < 0) {
		qdma_log_error(
		"%s Buff too small, err:%d\\n",
		__func__,
		-QDMA_ERR_NO_MEM);
		return -QDMA_ERR_NO_MEM;
	}
	len += rv;

"""

dump_function_code ="""
{
	uint32_t num_regs =
		sizeof(reg_dump_arr)/
		sizeof((reg_dump_arr)[0]);
	int i = 0;
	int rv = 0;

	for (i = 0; i < num_regs; i++) {
		if (reg_addr == reg_dump_arr[i].reg_address)
			rv = reg_dump_arr[i].reg_info_dump_fn(reg_val, buf, buflen);
	}
	return rv;
}


"""

def rchop(s, suffix):
    if suffix and s.endswith(suffix):
        return s[:-len(suffix)]
    return s


bitfield_macro_name_set = set()
# This python script generates .h and .c files from .ods file
def create_define(iptype, name, addr):
	addr_macro_name ='%s_%s_ADDR' %(iptype,name.upper())
	addr_macro_name = addr_macro_name.replace('__','_')
	macro_name='#define %-50s %s\n' %(addr_macro_name,addr)

	return macro_name

def create_bitfield(regname, bitfield_name, mask_details, long_bitfield):
	bitfield_name = bitfield_name.upper().replace('__','_')
	i=0

	bitfield_macro = regname+"_"+bitfield_name+"_MASK"
	
	while (True):
		if (bitfield_macro not in bitfield_macro_name_set):
			bitfield_macro_name_set.add(bitfield_macro)
			break
		else:
			i = i+1
		if (long_bitfield == False):
			bitfield_macro = regname+"_"+bitfield_name+"_"+str(i)+"_MASK"

	if (':' in mask_details):
		MSB = mask_details.split(':')[0].split('[')[1]
		LSB = mask_details.split(':')[1].split(']')[0]
		if (long_bitfield == False):
			macro_name='#define %-50s GENMASK(%s, %s)\n' %(bitfield_macro,MSB,LSB)
		else:
			MSB_d = int(MSB)
			LSB_d = int(LSB)

			if ( (MSB_d/32 == LSB_d/32) ):
				#The bitfield is not spread across muliple 4B registers
				regname = regname+"_W"+str(MSB_d/32)
				bitfield_macro = regname+"_"+bitfield_name+"_MASK"

				macro_name='#define %-50s GENMASK(%d, %d)\n' %(bitfield_macro,MSB_d%32,LSB_d%32)
			elif ((MSB_d - LSB_d < 32)):
				#The bitfield is spread across muliple 4B registers, but is less than 32 bits in width

				bitfield_macro = regname+"_W"+str(MSB_d/32)+"_"+bitfield_name+"_H_MASK"
				macro_name_h='#define %-50s GENMASK(%d, 0)\n' %(bitfield_macro,MSB_d%32)

				bitfield_macro = regname+"_W"+str(LSB_d/32)+"_"+bitfield_name+"_L_MASK"
				macro_name_l='#define %-50s GENMASK(31, %d)\n' %(bitfield_macro,LSB_d%32)

				macro_name = macro_name_h + macro_name_l
			else:
				bitfield_macro = regname+"_W"+str(MSB_d/32)+"_"+bitfield_name+"_H_MASK"
				macro_name_h='#define %-50s GENMASK(%d, 0)\n' %(bitfield_macro,MSB_d%32)

				macro_name_m = ""
				if(MSB_d/32 - LSB_d/32 > 1):
					bitfield_macro = regname+"_W"+str(MSB_d/32 - 1)+"_"+bitfield_name+"_M_MASK"
					macro_name_m='#define %-50s GENMASK(31, 0)\n' %(bitfield_macro)

				bitfield_macro = regname+"_W"+str(LSB_d/32)+"_"+bitfield_name+"_L_MASK"
				macro_name_l='#define %-50s GENMASK(31, %d)\n' %(bitfield_macro,LSB_d%32)
				macro_name = macro_name_h + macro_name_m + macro_name_l

	else:
		if (long_bitfield == False):
			BIT = mask_details.split('[')[1].split(']')[0]
			BIT_d = int(BIT) % 32
			macro_name='#define %-50s BIT(%d)\n' %(bitfield_macro,BIT_d)
		else:
			BIT = mask_details.split('[')[1].split(']')[0]
			BIT_d = int(BIT)
			bitfield_macro = regname+"_W"+str(BIT_d/32)+"_"+bitfield_name+"_MASK"
			macro_name='#define %-50s BIT(%d)\n' %(bitfield_macro,BIT_d%32)

	macro_name = macro_name.replace(' _MASK','_MASK')
	macro_name = macro_name.replace('_MASK_','_')
	macro_name = macro_name.replace('MASK_MASK','_MASK')
	macro_name = macro_name.replace('_C2H_STAT_','_')
	macro_name = macro_name.replace('QDMA_DBG_','')
	macro_name = macro_name.replace('_DBG_','_')
	macro_name = macro_name.replace('__','_')
	
	bitfield_macro = bitfield_macro.replace(' _MASK','_MASK')
	bitfield_macro = bitfield_macro.replace('_MASK_','_')
	bitfield_macro = bitfield_macro.replace('MASK_MASK','_MASK')
	bitfield_macro = bitfield_macro.replace('_C2H_STAT_','_')
	bitfield_macro = bitfield_macro.replace('QDMA_DBG_','')
	bitfield_macro = bitfield_macro.replace('_DBG_','_')
	bitfield_macro = bitfield_macro.replace('__','_')
	
	if (len(macro_name) > 80 ):
		macro_name.replace('\t\t','\t')
	return macro_name,bitfield_macro
	
def create_bitfield_table_struct ( bitfield_name):
	table_entry = '\n\t{\"%s\",\n\t\t%s},' %(rchop(bitfield_name, "_MASK"), bitfield_name)
	return table_entry

def create_register_name(reg_name_ip):
	reg_name = reg_name_ip.upper()
	reg_name = reg_name.split('[')[0]
	reg_name = reg_name.replace(' ','')
	reg_name = reg_name.replace('QDMA_','')
	reg_name = reg_name.replace('_MASK_','')
	reg_name = reg_name.replace('PERFORMANCE','PERF')
	reg_name = reg_name.replace('DEBUG','DBG')
	reg_name = reg_name.replace('MONITOR','MON')
	reg_name = reg_name.replace('REGISTER','REG')
	reg_name = reg_name.replace('COMMAND','CMD')
	reg_name = reg_name.replace('ERROR','ERR')
	reg_name = reg_name.replace('CONTROL','CTL')
	reg_name = reg_name.replace('FUNCTION','FUNC')
	reg_name = reg_name.replace('REQUEST','REQ')
	reg_name = reg_name.replace('PAYLOAD','PLD')
	reg_name = reg_name.replace('DESCRIPTOR','DESC')
	reg_name = reg_name.replace('COUNT','CNT')	
	reg_name = reg_name.replace('COMPLETED','CMPL')
	reg_name = reg_name.replace('CONFIG','CFG')
	reg_name = reg_name.replace('BLOCK','BLK')
	reg_name = reg_name.replace('CREDIT','CRED')
	reg_name = reg_name.replace('__','_')
	return reg_name

def create_bitmask_name(bitmask_name):
	#bitmask_name = sheet.cell(bitfield_row_index,BITFIELD_NAME_COL).value
	bitmask_name = bitmask_name.upper()
	bitmask_name = bitmask_name.strip()
	bitmask_name = bitmask_name.replace('\n','')
	bitmask_name = bitmask_name.replace(' ','')
	#bitmask_name = bitmask_name.replace('RESERVED','RSVD')
	bitmask_name = bitmask_name.replace('PERFORMANCE','PERF')
	bitmask_name = bitmask_name.replace('DEBUG','DBG')
	bitmask_name = bitmask_name.replace('MONITOR','MON')
	bitmask_name = bitmask_name.replace('REGISTER','REG')
	bitmask_name = bitmask_name.replace('COMMAND','CMD')
	bitmask_name = bitmask_name.replace('ERROR','ERR')
	bitmask_name = bitmask_name.replace('CONTROL','CTL')
	bitmask_name = bitmask_name.replace('FUNCTION','FUNC')
	bitmask_name = bitmask_name.replace('REQUEST','REQ')
	bitmask_name = bitmask_name.replace('PAYLOAD','PLD')
	bitmask_name = bitmask_name.replace('DESCRIPTOR','DESC')
	bitmask_name = bitmask_name.replace('COUNT','CNT')	
	bitmask_name = bitmask_name.replace('COMPLETED','CMPL')
	bitmask_name = bitmask_name.replace('CONFIG','CFG')
	bitmask_name = bitmask_name.replace('BLOCK','BLK')
	bitmask_name = bitmask_name.replace('CREDIT','CRED')
	bitmask_name = bitmask_name.replace('_C2H_STAT_','_')
	bitmask_name = bitmask_name.replace('_FATAL_','_')
	bitmask_name = bitmask_name.replace('QDMA_DBG_','')
	bitmask_name = bitmask_name.replace('_DBG_','_')
	return bitmask_name

script_desc="This python script generates .h and .c files from .ods file"
usage = """qdma_register_dump.py -f <ip .xlsx file> -i <ip type> -op_h <Output Header File> -op_c <Output .c file>"""
parser = ArgumentParser(prog='qdma_register_dump_generator.py', usage=usage, description=script_desc, add_help=True)

parser.add_argument("-f", "--odsfile",
                    action="store", type=str, dest="inputfile",
                    help = "The input ods file")
					
parser.add_argument("-i", "--iptype",
                    action="store", type=str, dest="iptype",
                    help = "IP Type")					

parser.add_argument("-op_h", "--op_header_file",
                    action="store", type=str, dest="op_h_file",
                    help = "Output Header File")

parser.add_argument("-op_c", "--op_source_code_file",
                    action="store", type=str, dest="op_c_file",
                    help = "Output Source Code File")

args = parser.parse_args()

inputfile = args.inputfile
iptype = args.iptype
output_header_file_name = args.op_h_file
output_src_code_file_name = args.op_c_file
if (inputfile == None or iptype == None or output_header_file_name == None or output_src_code_file_name == None ):
	print(usage)
	sys.exit(1)


output_header_file = open(output_header_file_name, "w") 
output_header_file.write(template_copyright)


output_header_file.write("#ifndef __"+output_header_file_name.upper().replace('.','_'))
output_header_file.write("\n#define __"+output_header_file_name.upper().replace('.','_')+"\n\n")


output_header_file.write(template_base_defines)

output_header_file.write("uint32_t "+iptype.lower()+"_config_num_regs_get(void);")
output_header_file.write("\nstruct xreg_info *"+iptype.lower()+"_config_regs_get(void);\n")

output_src_temp_file = open("temp_file", "w")

output_src_temp_file.write("static struct xreg_info "+iptype.lower()+"_config_regs[] = {\n")

output_src_code_file = open(output_src_code_file_name, "w")
output_src_code_file.write(template_copyright)
output_src_code_file.write("\n\n#include \""+output_header_file_name+"\"\n")
output_src_code_file.write("#include \"qdma_reg_dump.h\"\n\n")

windows_trace_enable_str = "#ifdef ENABLE_WPP_TRACING\n#include \""+output_src_code_file_name.lower().replace('.c','.tmh')+"\"\n#endif\n\n"

output_src_code_file.write(windows_trace_enable_str)

wb = openpyxl.load_workbook(inputfile)

#print (reg_name)
def parse_sheet(sheet_name, create_xregs_array, create_reg_define, create_bitfield_masks, create_dump_function):
	row_index = REGISTER_NAME_ROW_BASE
	sheet = wb[sheet_name]

	while (True):

		reg_name = sheet.cell(row_index,REGISTER_NAME_COL).value
		if ( reg_name == None):
			break;

		reg_name = reg_name.split('[')[0]
		reg_name = reg_name.strip()
		reg_name = reg_name.replace(' ','_')

		if (sheet_name == "Context_registers"):
			if(reg_name == "SOFTWARE_CONTEXT"):
				reg_name = "QDMA_SW_IND_CTXT_DATA_"
			elif(reg_name == "CREDIT_CONTEXT"):
				reg_name = "QDMA_CREDIT_CTXT_DATA_"
			elif(reg_name == "PREFETCH_CONTEXT"):
				reg_name = "QDMA_PREFETCH_CTXT_DATA_"
			elif(reg_name == "COMPLETION_CONTEXT"):
				reg_name = "QDMA_CMPL_CTXT_DATA_"
			elif(reg_name == "HARDWARE_CONTEXT"):
				reg_name = "QDMA_HW_IND_CTXT_DATA_"
			elif(reg_name == "INTERRUPT_CONTEXT"):
				reg_name = "QDMA_INTR_CTXT_DATA_"

		if (create_xregs_array == True):
				pf_vf_enabled = sheet.cell(row_index,PF_VF_COL).value
				featured_enabled = sheet.cell(row_index,FEATURE_COL).value

				if (iptype == "EQDMA"):
					debug_mode_enabled = sheet.cell(row_index,DEBUG_MODE_COL).value
					if (debug_mode_enabled == "C_DBG_EN"):
						debug_mode_enabled = "1"
					else :
						debug_mode_enabled = "0"
				else:
						debug_mode_enabled = "0"

				if (pf_vf_enabled == "PF"):
					pf_vf_enabled = "QDMA_REG_READ_PF_ONLY"
				elif (pf_vf_enabled == "VF"):
					pf_vf_enabled = "QDMA_REG_READ_VF_ONLY"
				elif (pf_vf_enabled == "PF_VF"):
					pf_vf_enabled = "QDMA_REG_READ_PF_VF"
				elif (pf_vf_enabled == "INVALID"):
					bitfield_row_index = row_index + 1
					while (sheet.cell(bitfield_row_index,BITFIELD_NAME_COL).value ):
						bitfield_row_index = bitfield_row_index + 1
					row_index = bitfield_row_index + 1
					if (str(sheet.cell(row_index,REGISTER_NAME_COL).value)[0] == '#' ) :
						row_index = row_index + 1
					
					continue
					
				else :
					print("Incorrectly populated sheet for PF/VF reg type"+str(row_index) + pf_vf_enabled )
					return
					

				if (featured_enabled == "MM"):
					featured_enabled = "QDMA_MM_MODE"
				elif (featured_enabled == "ST"):
					featured_enabled = "QDMA_ST_MODE"
				elif (featured_enabled == "MM_ST"):
					featured_enabled = "QDMA_MM_ST_MODE"
				elif (featured_enabled == "MBOX"):
					featured_enabled = "QDMA_MAILBOX"
				elif (featured_enabled == "CMPL"):
					featured_enabled = "QDMA_COMPLETION_MODE"
				else :
					print( "Incorrectly ppopulated sheet for MM/ST/MBOX etc")
					return

		reg_name = create_register_name(reg_name)

		reg_name_macro = create_define(iptype,reg_name, sheet.cell(row_index,REGISTER_ADDRESS_COL).value)
		
		bitfield_array_entry = "\nstatic struct regfield_info\n\t"+reg_name.lower()+"_field_info[] = {"

		if (create_xregs_array == True):			
			output_src_temp_file.write("{\""+reg_name + "\", "+sheet.cell(row_index,REGISTER_ADDRESS_COL).value.lower()+
			",\n\t1, 0, 0, 0,\n\t" + debug_mode_enabled+", "+featured_enabled+", "+pf_vf_enabled+ 
			",\n\tARRAY_SIZE("+reg_name.lower()+"_field_info),\n\t"+reg_name.lower()+"_field_info")
			output_src_code_file.write(bitfield_array_entry)
		reg_name = reg_name.strip()
		bitfield_row_index = row_index + 1
		#print ("Register address " + reg_name)
		if (create_reg_define == True):
			output_header_file.write(reg_name_macro)
		reserved_count = 0;
		while (sheet.cell(bitfield_row_index,BITFIELD_NAME_COL).value ):

			#print ("Bitmask name: " + sheet.cell(bitfield_row_index,BITFIELD_NAME_COL).value.upper())
			bitmask_name = sheet.cell(bitfield_row_index,BITFIELD_NAME_COL).value
			bitmask_name = create_bitmask_name(bitmask_name)

			common_str = SequenceMatcher(None, reg_name.upper(), bitmask_name ).find_longest_match(0, len(reg_name.upper()), 0, len(bitmask_name))

			if (common_str.size > 4):
				#print ("Register name: "+ reg_name.upper())
				#print ("Bitmask name: " + bitmask_name)

				bitmask_name = bitmask_name.replace(reg_name.upper()[common_str.a: common_str.a + common_str.size],'')
				#print ("updated bitmask   "+bitmask_name)
				#print ( "Common string is " + reg_name.upper()[common_str.a: common_str.a + common_str.size] )

			if (bitmask_name == "RESERVED" ):
				reserved_count = reserved_count + 1
				bitmask_name = "RSVD"
				bitmask_name = bitmask_name + "_"+str(reserved_count)

			#bitmask contains the #define that needs to be written to the .h file, bitmask_name is the name of the MACRO that gets defined
			if (sheet_name == "Context_registers"):
				bitmask,bitmask_name = create_bitfield(reg_name,bitmask_name,sheet.cell(bitfield_row_index, BITFIELD_DEFINITION_COL).value, True )
			else:
				bitmask,bitmask_name = create_bitfield(reg_name,bitmask_name,sheet.cell(bitfield_row_index, BITFIELD_DEFINITION_COL).value, False )
			output_header_file.write(bitmask)
			if (create_xregs_array == True):
				table_entry = create_bitfield_table_struct(bitmask_name)				
				output_src_code_file.write(table_entry)
					
			bitfield_row_index = bitfield_row_index + 1
		
			

		row_index = bitfield_row_index + 1
		if (create_dump_function == True):
				output_src_code_file.write("\n};\n\n")
		if (create_xregs_array == True):
			output_src_temp_file.write("\n},\n")
		num_blank_rows = 0
		while (num_blank_rows < 10):
			if (sheet.cell(row_index,REGISTER_NAME_COL).value):
				if (str(sheet.cell(row_index,REGISTER_NAME_COL).value)[0] == '#' ) :
					row_index = row_index + 1
				else:
					break
			else:
				row_index = row_index + 1
				num_blank_rows = num_blank_rows + 1


parse_sheet("qdma_regs",True,True,True,True)
parse_sheet("Context_registers",False,False,True,False)

output_header_file.write(template_footer)
output_header_file.close()

output_src_code_file.close()
output_src_temp_file.write("\n};" )
output_src_temp_file.write("\n\nuint32_t "+iptype.lower()+"_config_num_regs_get(void)\n{\n\treturn (sizeof("+iptype.lower()+"_config_regs)/\n\t\tsizeof("+iptype.lower()+"_config_regs[0]));\n}")
output_src_temp_file.write("\n\nstruct xreg_info *"+iptype.lower()+"_config_regs_get(void)\n{\n\treturn "+iptype.lower()+"_config_regs;\n}\n\n")
output_src_temp_file.close()
os.system("cat temp_file >> "+output_src_code_file_name)
